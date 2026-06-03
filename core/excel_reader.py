"""
core/excel_reader.py
--------------------
Read an Excel (.xlsx) file and return a list of row dictionaries.

Features:
  - First row is treated as headers
  - Empty cells become None
  - Date cells become DD/MM/YYYY
  - Whole-number floats become integer strings
  - Profile column mapping renames Excel headers to template/app fields
  - Duplicate headers are rejected because they make mapping ambiguous
"""

from datetime import date, datetime

import openpyxl


def _format_cell_value(value) -> str | None:
    """Convert an openpyxl cell value to a clean string or None."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, str):
        return value.strip() or None
    return str(value)


def _clean_headers(raw_headers) -> list[str]:
    return [
        str(header).strip() if header is not None else f"_col_{index}"
        for index, header in enumerate(raw_headers)
    ]


def _duplicate_headers(headers: list[str]) -> list[str]:
    seen: set[str] = set()
    duplicates: list[str] = []
    for header in headers:
        if header.startswith("_col_"):
            continue
        if header in seen and header not in duplicates:
            duplicates.append(header)
        seen.add(header)
    return duplicates


def _raise_duplicate_headers(headers: list[str]) -> None:
    duplicates = _duplicate_headers(headers)
    if duplicates:
        raise ValueError(
            "Excel file has duplicate column headers: "
            + ", ".join(duplicates)
            + ". Rename duplicate columns before loading this file."
        )


def read_excel(
    excel_path: str,
    column_mapping: dict | None = None,
    required_fields: list[str] | None = None,
    sheet_name: str | None = None,
) -> list[dict]:
    """
    Read an Excel file and return rows as dictionaries.

    Args:
        excel_path: Full path to the .xlsx file.
        column_mapping: Optional {output_key: excel_column_header} map.
        required_fields: Output keys whose mapped Excel columns must exist.
        sheet_name: Sheet to read. If None, reads the active sheet.
    """
    if not excel_path:
        raise ValueError("Excel path cannot be empty.")

    wb = None
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows_iter = ws.iter_rows(values_only=True)

        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            raise ValueError("Excel file is empty - no header row found.")

        headers = _clean_headers(raw_headers)
        if not any(header for header in headers if not header.startswith("_col_")):
            raise ValueError("Excel file header row appears empty.")
        _raise_duplicate_headers(headers)

        if column_mapping:
            missing_cols = [
                f"'{excel_col}' (needed for '{std_key}')"
                for std_key, excel_col in column_mapping.items()
                if excel_col not in headers and (not required_fields or std_key in required_fields)
            ]
            if missing_cols:
                raise ValueError(
                    "Excel file is missing required columns:\n"
                    + "\n".join(f"  - {column}" for column in missing_cols)
                    + f"\n\nAvailable columns: {headers}"
                )

        result: list[dict] = []
        for raw_row in rows_iter:
            raw_row_dict: dict[str, str | None] = {}
            for header, cell_value in zip(headers, raw_row):
                raw_row_dict[header] = _format_cell_value(cell_value)

            row_dict: dict[str, str | None] = dict(raw_row_dict)
            if column_mapping:
                # Allow one Excel column to feed multiple template/app fields.
                for output_key, excel_col in column_mapping.items():
                    row_dict[output_key] = raw_row_dict.get(excel_col)

            if all(value is None for value in row_dict.values()):
                continue

            result.append(row_dict)

        return result
    finally:
        if wb is not None:
            wb.close()


def get_column_headers(excel_path: str, sheet_name: str | None = None) -> list[str]:
    """
    Return only the column headers from an Excel file.

    Used by the Profiles tab to populate mapping dropdowns.
    """
    wb = None
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        header_row = next(ws.iter_rows(values_only=True), None)
        if header_row is None:
            return []
        headers = [str(header).strip() for header in header_row if header is not None]
        _raise_duplicate_headers(headers)
        return headers
    finally:
        if wb is not None:
            wb.close()
